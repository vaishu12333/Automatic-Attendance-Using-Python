import cv2
import numpy as np
import pyttsx3
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import os

# ---------------- CAMERA SETUP ----------------
# External camera usually index 1
print("Program started")

cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
print("Trying camera...")

if not cap.isOpened():
    print("Camera not detected")
    exit()

print("Camera opened")
detector = cv2.QRCodeDetector()

# ---------------- STUDENT DATA ----------------
rollnop = []

roll = {
1: "Ayushi Bhandari",
2: "Gayatri Chaudhari",
3: "Srushti Chaudhary",
4: "Sujal Deore",
5: "Pawan Desale",
6: "Hemant Deshmukh",
7: "Vaishnavi Dusane",
8: "Naitik Fulfagar",
9: "Saee Gadhekar",
10: "Payal Gaikwad",
11: "Vaibhavi Gangurde",
12: "Aditya Jadhav",
13: "Rahul Jadhav",
14: "Ujjwal Katare",
15: "Cancelled",
16: "Priti Khaire",
17: "Tanmay Khairnar",
18: "Rutuja Lomate",
19: "Ashwini Mali",
20: "Amey Mohole",
21: "Parth Pagar",
22: "Durgesh Pagar",
23: "Keshavdas Panpatil",
24: "Mayuri Pardeshi",
25: "Ritesh Pardeshi",
26: "Tanvi Pardeshi",
27: "Tejas Patil",
28: "Vishnavi Patil",
29: "Tilak Rokaya",
30: "Sudarshan Sanap",
31: "Taha Shaikh",
32: "Ajinkya Shinde",
33: "Sayee Shirsath",
34: "Aditya Sonawane",
35: "Sumedh Takalkar",
36: "Varad Tannu",
37: "Pranav Thorat",
38: "Aditya Tokare",
39: "Rutuja Vaidya",
40: "Aashish Wagh",
41: "Prajwal Wankhede",
42: "Prerana Patil",
43: "Ruchika Kapadnis"
}

# ---------------- VOICE ----------------
engine = pyttsx3.init()
engine.setProperty('rate', 150)
engine.setProperty('volume', 1.0)

# ---------------- DATE FOLDER ----------------
x = datetime.now()
dirName = x.strftime("%d-%m-%Y")

if not os.path.exists(dirName):
    os.mkdir(dirName)

# ---------------- EXCEL FILE ----------------
file_path = dirName + "/Attendance.xlsx"

wb = openpyxl.Workbook()
sheet = wb.active

sheet['A1'] = "Roll No"
sheet['B1'] = "Name"
sheet['C1'] = "In Time"
sheet['D1'] = x.strftime("%d-%m-%Y")
sheet['E1'] = "Present Count"

for i in range(1, 44):
    sheet.cell(row=i+1, column=1).value = i
    sheet.cell(row=i+1, column=2).value = roll[i]

wb.save(file_path)

# ---------------- QR SCAN FUNCTION ----------------
def decoder(image):
    data, bbox, _ = detector.detectAndDecode(image)

    if data:
        try:
            barcodeDataint = int(data)

            if barcodeDataint in roll and barcodeDataint not in rollnop:

                name = roll[barcodeDataint]
                print("Present:", name)

                engine.say(name + " Present")
                engine.runAndWait()

                rollnop.append(barcodeDataint)

                workbook = load_workbook(file_path)
                sheet = workbook.active

                now = datetime.now().strftime("%H:%M:%S")

                sheet.cell(row=barcodeDataint+1, column=3).value = now
                sheet['E2'] = len(rollnop)

                workbook.save(file_path)

            elif barcodeDataint in rollnop:
                print("Duplicate Entry")
                engine.say("Duplicate Entry")
                engine.runAndWait()

            else:
                print("Invalid QR")
                engine.say("Invalid Input")
                engine.runAndWait()

        except:
            pass

# ---------------- MAIN LOOP ----------------
while True:
    ret, frame = cap.read()

    if not ret:
        print("Camera not detected")
        break

    decoder(frame)

    cv2.imshow("Attendance Scanner", frame)

    if cv2.waitKey(1) == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()